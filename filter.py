# -*- coding: utf-8 -*-
import xlrd
import openpyxl
import re
import os
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showerror, showinfo


def get_data_path():
    re_data_path = re.compile(r'^[1-9]{1}[0-9]{3}.xlsx$')   

    data_path = []
    for path in os.listdir():
        if re_data_path.match(path):
            data_path.append(path)
    if len(data_path) == 0:
        showerror(title='error', message='could not find data file')
    return data_path


def search(key, none_value):
    data_path = get_data_path()
    # print(data_path)

    # key = 106001
    # none_value = 1

    time = []
    sale_count = []
    sale_money = []
    people = []
    for path in data_path:
        data = xlrd.open_workbook(path)

        for i in range(1, 13):
            table = data.sheet_by_name(f'{i}月')
            for j in range(3-1, table.nrows-1):
                # (行，列)
                if key == table.cell(j, 0).value:
                    a = table.cell(j, 4).value if table.cell(j, 4).value else none_value
                    b = table.cell(j, 5).value if table.cell(j, 5).value else none_value
                    c = table.cell(j, 6).value if table.cell(j, 6).value else none_value
                    time.append(f'{path[0:4]}-{i}')
                    sale_count.append(a)
                    sale_money.append(b)
                    people.append(c)
    if len(time) == 0:
        showinfo(title='info', message='does not have any data')
        return
    draw(time, [
        {'name': 'sale_count', 'value': sale_count},
        {'name': 'sale_money', 'value': sale_money},
        {'name': 'people', 'value': people}])


def calculate(none_value, save_data):
    data_path = get_data_path()
    time = []
    month_total_sale_count = []
    month_total_sale_money = []
    month_total_people = []
    for path in data_path:
        data = xlrd.open_workbook(path)
        for i in range(1, 13):
            time.append(f'{path[0:4]}-{i}')     
            table = data.sheet_by_name(f'{i}月')
            month_total_sale_count.append(sum(value if value else none_value for value in table.col_values(4)[2:table.nrows-1]))       
            month_total_sale_money.append(sum(value if value else none_value for value in table.col_values(5)[2:table.nrows-1]))
            month_total_people.append(sum(value if value else none_value for value in table.col_values(6)[2:table.nrows-1]))
    l = len(time)
    month_avg_sale_count = list(map(lambda x:x/l, month_total_sale_count))
    month_avg_sale_money = list(map(lambda x:x/l, month_total_sale_money))
    month_avg_people = list(map(lambda x:x/l, month_total_people))
    if len(time) == 0:
        showinfo(title='info', message='does not have any data')
        return
    data = [
        {'name': 'month_total_sale_count', 'value': month_total_sale_count},
        {'name': 'month_total_sale_money', 'value': month_total_sale_money},
        {'name': 'month_total_people', 'value': month_total_people},
        {'name': 'month_avg_sale_count', 'value': month_avg_sale_count},
        {'name': 'month_avg_sale_money', 'value': month_avg_sale_money},
        {'name': 'month_avg_people', 'value': month_avg_people},
    ]
    if save_data == 1:
        save(time, data)
    draw(time, data)


def draw(time, draw_list):
    for j, item in enumerate(draw_list):        
        plt.figure(j)
        plt.subplot(1, 1, 1)
        plt.title(item['name'])
        plt.plot(time, item['value'], 'o-')
        plt.xlabel('time')
        plt.ylabel(item['name'])
        for i in range(1, len(time), 3):
            plt.text(time[i], item['value'][i], item['value'][i], ha='right',
                     va='bottom', fontsize=10)
        plt.gcf().autofmt_xdate()  # 自动旋转日期标记
    plt.show()


def save(time, data):
    book = openpyxl.Workbook()  # 新建工作簿
    book.create_sheet('total')  # 添加页
    # table = data.get_sheet_by_name('Sheet1') # 获得指定名称页
    table = book.active  # 获得当前活跃的工作页，默认为第一个工作页
    items = []
    for j in range(len(data)):
        table.cell(j+2, 1, data[j]['name'])
        items.append(data[j]['value'])
    for i in range(len(time)):
        table.cell(1, i+2, time[i])  # 行，列，值 这里是从1开始计数的
        for k, item in enumerate(items):
            table.cell(k+2, i+2, item[i])
    book.save('calculate_result.xlsx')  # 一定要保存


def gui():
    try:
        data = xlrd.open_workbook('./category.xlsx')
    except FileNotFoundError:
        showerror(title='error', message='could not find ./category.xlsx')
        return
    table = data.sheets()[0]
    category = table.col_values(0)[2:table.nrows-1]

    window = tk.Tk()
    window.title('filter data')
    window.geometry('500x300')

    ttk.Label(window, text='none_value').grid(row=1, column=1)
    none_value = tk.Entry(window, show=None, font=('Arial', 14))
    none_value.grid(row=1, column=2)
    
    ttk.Label(window, text='category').grid(row=3, column=1)

    key_values = tk.StringVar()  # 窗体自带的文本，新建一个值
    key = ttk.Combobox(window, textvariable=key_values)  # 初始化
    key["values"] = category
    key.grid(row=3, column=2)

    ttk.Label(window, text='total calculate').grid(row=4, column=1)
    c = tk.IntVar()
    c.set(1)
    check = tk.Checkbutton(
        window, text='save data', variable=c, onvalue=1, offvalue=2)
    check.grid(row=4, column=2)

    def trigger_search():
        if not key.get():
            showerror(title='error', message='you should select category')
            return
        if not none_value.get():
            showerror(title='error', message='you should input none_value')
            return
        search(int(float(key.get())), float(none_value.get()))

    def trigger_calculate():
        if not none_value.get():
            showerror(title='error', message='you should input none_value')
            return
        calculate(int(float(none_value.get())), c.get())

    tk.Button(window, text='search', width=10, height=2,
              command=trigger_search).grid(row=3, column=4)
    tk.Button(window, text='calculate', width=10, height=2,        
              command=trigger_calculate).grid(row=4, column=4)
    # 进入消息循环
    window.mainloop()


if __name__ == '__main__':
    gui()


    


